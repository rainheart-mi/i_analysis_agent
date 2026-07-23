"""
AMIS 导出服务：解析 artifact_schema 中的表格结构，从 artifact_data 提取数据，
生成 XLSX 文件。

支持两种数据类型：
  1. n8n 节点：schema 中有 crud/table 组件，data 由 source 路径引用
  2. agent 节点：schema 是自包含 page，data 在 schema 的 data 字段中
"""
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

# AMIS schema 递归遍历的 key
_RECURSIVE_KEYS = ("body", "columns", "items", "tabs", "header", "footer", "toolbar")


def _resolve_path(data: dict, path: str) -> Any | None:
    """解析模板路径（如 `${processedData.data}`）到实际数据。"""
    if not path:
        return None
    # 去掉 ${} 包裹
    inner = path.strip()
    if inner.startswith("${") and inner.endswith("}"):
        inner = inner[2:-1]
    # 按 . 分割路径
    parts = inner.split(".")
    current = data
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list) and part.isdigit():
            current = current[int(part)]
        else:
            return None
    return current


def _find_tables_in_schema(schema: dict) -> list[dict]:
    """递归遍历 AMIS schema，找出所有 crud/table 组件。

    Returns:
      [{"title": str, "source": str, "columns": [{"label", "name"}], "data": [...]}, ...]
    """
    tables = []
    tabs_title = ""  # 当前 tabs 的 title

    def _walk(node: dict, tab_title: str = ""):
        if not isinstance(node, dict):
            return
        t = node.get("type")
        # 记录当前 tab 标题
        if t == "tabs" and "tabs" in node:
            for tab in node.get("tabs", []):
                _walk(tab, tab.get("title", ""))

        # 找到 crud 或 table 组件
        if t in ("crud", "table"):
            source = node.get("source", "")
            columns = [
                {"label": c.get("label", ""), "name": c.get("name", "")}
                for c in node.get("columns", [])
            ]
            title = node.get("title", node.get("name", ""))
            # 尝试从 source 解析数据
            prepared: list[dict] = []
            # 先标记，后面再填充 data
            tables.append({
                "title": title or tab_title or "Sheet1",
                "source": source,
                "columns": columns,
                "data": [],
                "_tab_title": tab_title,
            })
            return

        # 递归遍历
        for key in _RECURSIVE_KEYS:
            child = node.get(key)
            if isinstance(child, list):
                for item in child:
                    if isinstance(item, dict):
                        _walk(item, tab_title)
            elif isinstance(child, dict):
                _walk(child, tab_title)

        # 处理 tabs 内的子项
        if t == "tabs" and "tabs" in node:
            return

    _walk(schema)
    return tables


def _extract_data_from_agent_schema(schema: dict) -> list[dict]:
    """从 agent 自包含 schema 提取表格数据。

    agent 的 schema 是 page 类型，data 在 schema.data 字段中。
    把 data 的嵌套结构拍平为表格。
    """
    data = schema.get("data", {})
    if not data:
        return []

    tables = []
    # 处理 categories 表
    categories = data.get("categories", [])
    if categories:
        tables.append({
            "title": "品类总览",
            "source": "data.categories",
            "columns": [
                {"label": "品类名称", "name": "name"},
                {"label": "跨度", "name": "colspan"},
            ],
            "data": categories,
        })

    # 处理 bands 表（价格带+商品明细）
    bands = data.get("bands", [])
    if bands:
        sub_headers = data.get("subHeaders", [])
        rows_data = []
        for band in bands:
            label = band.get("label", "")
            for row in band.get("rows", []):
                cells = row.get("cells", [])
                for cell in cells:
                    product = cell.get("product", {})
                    row_data = {"价格带": label}
                    # 如果 subHeaders 存在，按顺序映射
                    if sub_headers:
                        header_map = {
                            "品名": "name",
                            "售价": "price",
                            "ABC": "abc",
                            "PSD": "psd",
                            "销量": "soldQty",
                            "销售额": "revenue",
                            "天数": "days",
                        }
                        for h in sub_headers:
                            key = header_map.get(h, h)
                            row_data[h] = product.get(key, "")
                    else:
                        row_data.update(product)
                    rows_data.append(row_data)

        if rows_data:
            tables.append({
                "title": "商品明细",
                "source": "data.bands",
                "columns": [
                    {"label": k, "name": k} for k in rows_data[0].keys()
                ],
                "data": rows_data,
            })

    return tables


def generate_xlsx(artifact_data: dict, artifact_schema: dict) -> io.BytesIO:
    """根据 artifact_data 和 artifact_schema 生成 XLSX 文件。

    Args:
        artifact_data: 节点执行结果数据（含 processedData 等）
        artifact_schema: AMIS schema 定义（含 crud/table 组件）

    Returns:
        XLSX 文件的 BytesIO 流
    """
    wb = Workbook()
    # 删除默认空白 sheet
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. 从 schema 中找表格结构
    tables = _find_tables_in_schema(artifact_schema)

    # 2. 如果是 agent 自包含 schema，也用 schema.data 解析
    if not tables and artifact_schema.get("type") == "page" and artifact_schema.get("data"):
        tables = _extract_data_from_agent_schema(artifact_schema)

    # 3. 如果是 n8n 数据，仍然没找到表格，fallback 到处理整个 artifact_data
    if not tables:
        # 尝试从 processedData 提取数组
        pd = artifact_data.get("processedData", artifact_data)
        # 找第一个数组
        for key, value in pd.items():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                tables.append({
                    "title": key,
                    "source": f"processedData.{key}",
                    "columns": [{"label": k, "name": k} for k in value[0].keys()],
                    "data": value,
                })
                break

    # 4. 填充数据并生成 sheet
    for table in tables:
        # 如果 data 还没填充，从 source 解析
        if not table["data"]:
            raw_data = _resolve_path(artifact_data, table["source"])
            if isinstance(raw_data, list):
                table["data"] = raw_data

        data = table["data"]
        if not data:
            continue

        # sheet 名最大 31 字符
        sheet_name = table["title"][:31]
        ws = wb.create_sheet(title=sheet_name)

        # 列定义：优先用 schema 的 columns，否则从数据第一行推断
        columns = table["columns"]
        if not columns and data:
            columns = [{"label": k, "name": k} for k in data[0].keys()]

        # 写表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["label"])
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 写数据行
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                name = col["name"]
                value = row_data.get(name) if isinstance(row_data, dict) else None
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # 数字右对齐
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # 自动列宽（近似）
        for col_idx, col in enumerate(columns, 1):
            max_len = len(col["label"])
            for row_idx in range(2, len(data) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            # 设置列宽（中文大约 2 倍宽度）
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_len * 2, 60)

    # 处理 edge case：没有找到任何表格数据
    if wb.worksheets:
        # 调整所有 sheet
        pass
    else:
        ws = wb.create_sheet(title="数据")
        ws.cell(row=1, column=1, value="未找到可导出的表格数据")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output